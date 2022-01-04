import openpyxl
import sys
import os

def main():
    inputname = get_txt_list()
    outputname = get_xlsx_lsit()
    for iin, oout in zip(inputname, outputname):
        print(iin, oout)
        convert(iin,oout)

def get_txt_list():
    path = "./"
    file_list = os.listdir(path)
    file_list_txt = [file for file in file_list if file.endswith(".txt")]
    return file_list_txt

def get_xlsx_lsit():
    list = get_txt_list()
    new_strings = []
    for string in list:
        new_string = string.replace('.txt','.xlsx')
        new_strings.append(new_string)
    return new_strings

def convert(inputname,outputname):
    f = open(inputname, "r", encoding='utf-8')
    list = f.readlines()
    f.close()

    wb = openpyxl.Workbook()
    wb.save(outputname)

    file = openpyxl.load_workbook(outputname)
    sheet = file.active

    for i in range(1, len(list)+1):
        list1 = list[i-1].split('|')
        for k in range(1,len(list1)+1):
            sheet[chr(k+64)+str(i)].value = list1[k-1]
    file.save(outputname)

def errExit(msg):
    sys.stderr.write(msg + '\n')
    sys.exit(0)

if __name__ == '__main__':
    main()

# f = open('new_01_JMTjokbal.txt', "r", encoding='utf-8')
# list = f.readlines()
# f.close()
#
# wb = openpyxl.Workbook()
# wb.save('new_01_JMTjokbal.xlsx')
#
# file = openpyxl.load_workbook('new_01_JMTjokbal.xlsx')
# sheet = file.active
#
# for i in range(1, len(list) + 1):
#     list1 = list[i - 1].split('|')
#     for k in range(1, len(list1) + 1):
#         sheet[chr(k + 64) + str(i)].value = list1[k - 1]
# file.save('new_01_JMTjokbal.xlsx')