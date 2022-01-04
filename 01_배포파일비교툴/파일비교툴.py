# -*- coding: utf-8 -*-
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtWidgets import QApplication, QProgressBar
import os
import codecs
import chardet
import openpyxl
import time
import csv
import pandas as pd
from dbfread import DBF
from detect_delimiter import detect
import xlrd
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import  FormulaRule

class Ui_Folder_Compare(object):
    def setupUi(self, Folder_Compare):
        Folder_Compare.setObjectName("Folder_Compare")
        Folder_Compare.resize(813, 241)
        self.widget = QtWidgets.QWidget(Folder_Compare)
        self.widget.setGeometry(QtCore.QRect(6, 10, 801, 221))
        self.widget.setObjectName("widget")
        self.gridLayout = QtWidgets.QGridLayout(self.widget)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.gridLayout.setObjectName("gridLayout")
        self.label = QtWidgets.QLabel(self.widget)
        self.label.setObjectName("label")
        self.gridLayout.addWidget(self.label, 0, 0, 1, 1)
        self.Standard_line = QtWidgets.QLineEdit(self.widget)
        self.Standard_line.setObjectName("Standard_line")
        self.gridLayout.addWidget(self.Standard_line, 0, 1, 1, 1)
        self.Start_bt = QtWidgets.QPushButton(self.widget)
        self.Start_bt.setObjectName("Start_bt")
        self.gridLayout.addWidget(self.Start_bt, 0, 2, 1, 1)
        self.label_2 = QtWidgets.QLabel(self.widget)
        self.label_2.setObjectName("label_2")
        self.gridLayout.addWidget(self.label_2, 1, 0, 1, 1)
        self.Compare_line = QtWidgets.QLineEdit(self.widget)
        self.Compare_line.setObjectName("Compare_line")
        self.gridLayout.addWidget(self.Compare_line, 1, 1, 1, 1)
        self.End_bt = QtWidgets.QPushButton(self.widget)
        self.End_bt.setObjectName("End_bt")
        self.gridLayout.addWidget(self.End_bt, 1, 2, 1, 1)
        self.label_3 = QtWidgets.QLabel(self.widget)
        self.label_3.setObjectName("label_3")
        self.gridLayout.addWidget(self.label_3, 2, 0, 1, 1)
        self.progressBar = QtWidgets.QProgressBar(self.widget)
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        self.gridLayout.addWidget(self.progressBar, 2, 1, 1, 1)
        self.textBrowser = QtWidgets.QTextBrowser(self.widget)
        self.textBrowser.setObjectName("textBrowser")
        self.gridLayout.addWidget(self.textBrowser, 3, 0, 1, 3)

        self.retranslateUi(Folder_Compare)
        QtCore.QMetaObject.connectSlotsByName(Folder_Compare)

        self.Start_bt.clicked.connect(self.Compare_Main)
        self.End_bt.clicked.connect(QCoreApplication.instance().quit)

        style ="""
        QProgressBar {
            border: 2px solid grey;
            border-radius: 5px;
            text-align: center;
        }"""
        style2 ="""
        QLineEdit {
            border: 2px solid grey;
            border-radius: 5px;
            text-align: center;
        }"""
        style3 ="""
        QPushButton {
            border: 2px solid grey;
            border-radius: 2px;
            text-align: center;
        }"""
        style4 ="""
        QTextBrowser {
            border: 2px solid grey;
            border-radius: 2px;
            text-align: center;
        }"""
        self.progressBar.setStyleSheet(style)
        self.Standard_line.setStyleSheet(style2)
        self.Compare_line.setStyleSheet(style2)
        self.Start_bt.setStyleSheet(style3)
        self.End_bt.setStyleSheet(style3)
        self.textBrowser.setStyleSheet(style4)

    def retranslateUi(self, Folder_Compare):
        _translate = QtCore.QCoreApplication.translate
        Folder_Compare.setWindowTitle(_translate("Folder_Compare", "Folder_Compare_tool"))
        self.label.setText(_translate("Folder_Compare", "기준경로"))
        self.label.setFont(QtGui.QFont('기준경로',9, QtGui.QFont.Bold))
        self.Start_bt.setText(_translate("Folder_Compare", "시 작"))
        self.Start_bt.setFont(QtGui.QFont('시 작', 9, QtGui.QFont.Bold))
        self.label_2.setText(_translate("Folder_Compare", "비교경로"))
        self.label_2.setFont(QtGui.QFont('비교경로', 9, QtGui.QFont.Bold))
        self.End_bt.setText(_translate("Folder_Compare", "종 료"))
        self.End_bt.setFont(QtGui.QFont('종 료', 9, QtGui.QFont.Bold))
        self.label_3.setText(_translate("Folder_Compare", "진행도"))
        self.label_3.setFont(QtGui.QFont('진행도', 9, QtGui.QFont.Bold))

    def file_remove(self):
        file1 = '기준_파일리스트.txt'
        file2 = '비교_파일리스트.txt'
        if os.path.isfile(file1):
            os.remove(file1)
            print('기준_파일리스트 파일 삭제 완료')
        if os.path.isfile(file2):
            os.remove(file2)
            print('비교_파일리스트 파일 삭제 완료')

    def stand_list_save(self,dirname):
        dirname_fix = r'{}'.format(dirname)
        outfile = codecs.open('기준_파일리스트.txt', 'a')
        file_list = []
        filenames = os.listdir(dirname_fix)
        for filename in filenames:
            full_filename = os.path.join(dirname_fix, filename)
            if os.path.isdir(full_filename):
                self.stand_list_save(full_filename)
            else:
                # ext = os.path.splitext(full_filename)[-1]
                # if ext == '.csv':
                print(full_filename)
                fullname_fix = full_filename.upper()
                fullname = full_filename + '\n'
                if fullname_fix.endswith('.BMP') or fullname_fix.endswith('.MAX') or fullname_fix.endswith(
                    '.PVR') or fullname_fix.endswith('.TBM') or fullname_fix.endswith('.JPG') or fullname_fix.endswith(
                    '.PNG') or fullname_fix.endswith('.POD') or fullname_fix.endswith('THUMBS.DB') or fullname_fix.endswith('.SHX')== True:
                    pass
                else:
                    outfile.write(fullname)
                    file_list.append(full_filename)
                    # QApplication.processEvents()
        outfile.close()
        return file_list

    def compare_list_save(self,dirname):
        dirname_fix = r'{}'.format(dirname)
        outfile = codecs.open('비교_파일리스트.txt', 'a')
        file_list = []
        filenames = os.listdir(dirname_fix)
        for filename in filenames:
            full_filename = os.path.join(dirname_fix, filename)
            if os.path.isdir(full_filename):
                self.compare_list_save(full_filename)
            else:
                # ext = os.path.splitext(full_filename)[-1]
                # if ext == '.csv':
                print(full_filename)
                fullname_fix = full_filename.upper()
                fullname = full_filename + '\n'
                if fullname_fix.endswith('.BMP') or fullname_fix.endswith('.MAX') or fullname_fix.endswith(
                    '.PVR') or fullname_fix.endswith('.TBM') or fullname_fix.endswith('.JPG') or fullname_fix.endswith(
                    '.PNG') or fullname_fix.endswith('.POD') or fullname_fix.endswith('THUMBS.DB') or fullname_fix.endswith('.SHX') == True:
                    pass
                else:
                    outfile.write(fullname)
                    file_list.append(full_filename)
                    # QApplication.processEvents()
        outfile.close()
        return file_list

    def stand_list(self):
        with open('기준_파일리스트.txt') as data:
            lines = data.read().splitlines()
        SearchList = lines
        return SearchList

    def compare_list(self):
        with open('비교_파일리스트.txt') as data:
            lines = data.read().splitlines()
        SearchList = lines
        return SearchList

    def list_compare(self):
        stand_text = self.Standard_line.text()
        compare_text = self.Compare_line.text()
        A_list_fix_name = r'{}'.format(stand_text)
        B_list_fix_name = r'{}'.format(compare_text)

        A_list_fix = []
        for i in self.stand_list():
            s = i.split(stand_text)[1]
            A_list_fix.append(s)
        B_list_fix = []
        for i in self.compare_list():
            s = i.split(compare_text)[1]
            B_list_fix.append(s)

        # 두리스트 간의 차집합 리스트 추출
        A_list = set(A_list_fix)
        B_list = set(B_list_fix)

        A_only = A_list - B_list
        A_only_list = []
        for i in A_only:
            s = (A_list_fix_name + '\\' + i)
            A_only_list.append(s)

        B_only = B_list - A_list
        B_only_list = []
        for i in B_only:
            s = (B_list_fix_name + '\\' + i)
            B_only_list.append(s)

        # 두 리스트 간의 교집합 추출
        Intersection = A_list & B_list
        Intersection_list = []
        for i in Intersection:
            a = (A_list_fix_name + '\\' + i)
            b = (B_list_fix_name + '\\' + i)
            Intersection_list.append({'a': a, 'b': b})
        return A_only_list, B_only_list, Intersection_list

    def detail_compare(self):
        # 폴더 비교 결과 엑셀 파일 형식 지정
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'COMPARE'
        sheet.append(
            ["FILE_NAME", "STAN_CAPACITY", "COMPARE_CAPACITY", "DIFF_CAPACITY", "HEADER", "STAN_COUNT",
             "COMPARE_COUNT", "DIFF_COUNT", "ENCODING", "DELIMITER", "STAN_FOLDER", "COMPARE_FOLDER"])

        # 차,교집합 리스트 들 소환
        A_only_list = self.list_compare()[0]
        B_only_list = self.list_compare()[1]
        Intersection_list = self.list_compare()[2]

        # 차집합 리스트 엑셀 우선 저장
        for s in A_only_list:
            FILE_NAME = os.path.basename(s)
            sheet.append([FILE_NAME, '', '', '', '', '', '', '', '', '', s, ''])

        for s in B_only_list:
            FILE_NAME = os.path.basename(s)
            sheet.append([FILE_NAME, '', '', '', '', '', '', '', '', '', '', s])

        # 비교 함수 본체
        self.progressBar.setMaximum(len(Intersection_list))
        count = 0
        for list in Intersection_list:
            count+=1
            self.progressBar.setValue(count)
            A_Intersection_list = list['a']
            B_Intersection_list = list['b']
            print(A_Intersection_list, '   @@@@   ', B_Intersection_list)

            # 1.파일 이름 만 저장
            FILE_NAME = os.path.basename(A_Intersection_list)
            # 1. 용량 추출
            A_Intersection_list_capacity = round(os.path.getsize(A_Intersection_list) / (1024 * 1024), 2)
            B_Intersection_list_capacity = round(os.path.getsize(B_Intersection_list) / (1024 * 1024), 2)
            # 2. B-A 용량
            DIFF_CAPACITY = round(B_Intersection_list_capacity - A_Intersection_list_capacity, 2)

            # 3. HEADER 비교
            HEADER = '-'
            # 파일이 csv 일경우
            if A_Intersection_list.endswith('.csv') == True and B_Intersection_list.endswith('.csv') == True:
                try:
                    A = open(A_Intersection_list)
                    B = open(B_Intersection_list)
                    A_HEADER = next(csv.reader(A))
                    B_HEADER = next(csv.reader(B))
                except:
                    A = open(A_Intersection_list, encoding='utf-8')
                    B = open(B_Intersection_list, encoding='utf-8')
                    A_HEADER = next(csv.reader(A))
                    B_HEADER = next(csv.reader(B))
                # print(A_HEADER,'//',B_HEADER)
                HEADER = str(str(A_HEADER) == str(B_HEADER))
            # 파일이 xlsx 일경우
            if A_Intersection_list.endswith('.xlsx') == True and B_Intersection_list.endswith('.xlsx') == True:
                A = pd.read_excel(A_Intersection_list)
                B = pd.read_excel(B_Intersection_list)
                A_HEADER = A.columns
                B_HEADER = B.columns
                # print(A_HEADER,'//',B_HEADER)
                HEADER = str(str(A_HEADER) == str(B_HEADER))
            # 파일이 xls 일경우
            if A_Intersection_list.endswith('.xls') == True and B_Intersection_list.endswith('.xls') == True:
                A = pd.read_excel(A_Intersection_list)
                B = pd.read_excel(B_Intersection_list)
                A_HEADER = A.columns
                B_HEADER = B.columns
                # print(A_HEADER, '//', B_HEADER)
                HEADER = str(str(A_HEADER) == str(B_HEADER))
            # shp 일 경우 dbf파일의 헤더를 비교
            if A_Intersection_list.endswith('.shp') == True and B_Intersection_list.endswith('.shp') == True:
                try:
                    AA = A_Intersection_list.replace('.shp', '.dbf')
                    A = DBF(AA)
                    A_HEADER = A.field_names
                except:
                    AA = A_Intersection_list.replace('.shp', '.dbf')
                    A = DBF(AA, encoding='utf-8')
                    A_HEADER = A.field_names
                try:
                    BB = B_Intersection_list.replace('.shp', '.dbf')
                    B = DBF(BB)
                    B_HEADER = B.field_names
                except:
                    BB = B_Intersection_list.replace('.shp', '.dbf')
                    B = DBF(BB, encoding='utf-8')
                    B_HEADER = B.field_names
                # print(A_HEADER, '//', B_HEADER)
                HEADER = str(str(A_HEADER) == str(B_HEADER))
            # txt 파일 일 경우
            if A_Intersection_list.endswith('.txt') == True and B_Intersection_list.endswith('.txt') == True:
                try:
                    with open(A_Intersection_list) as data:
                        A_HEADER = data.readline()
                except:
                    try:
                        with open(A_Intersection_list, encoding='utf-8') as data:
                            A_HEADER = data.readline()
                    except:
                        with open(A_Intersection_list, encoding='ISO-8859-1') as data:
                            A_HEADER = data.readline()
                try:
                    with open(B_Intersection_list) as data:
                        B_HEADER = data.readline()
                except:
                    try:
                        with open(B_Intersection_list, encoding='utf-8') as data:
                            B_HEADER = data.readline()
                    except:
                        with open(B_Intersection_list, encoding='ISO-8859-1') as data:
                            B_HEADER = data.readline()
                # print(A_HEADER,'/',B_HEADER)
                HEADER = str(str(A_HEADER) == str(B_HEADER))
            # text 파일인 경우
            if A_Intersection_list.endswith('.text') == True and B_Intersection_list.endswith('.text') == True:
                try:
                    with open(A_Intersection_list) as data:
                        A_HEADER = data.readline()
                except:
                    try:
                        with open(A_Intersection_list, encoding='utf-8') as data:
                            A_HEADER = data.readline()
                    except:
                        with open(A_Intersection_list, encoding='ISO-8859-1') as data:
                            A_HEADER = data.readline()
                try:
                    with open(B_Intersection_list) as data:
                        B_HEADER = data.readline()
                except:
                    try:
                        with open(B_Intersection_list, encoding='utf-8') as data:
                            B_HEADER = data.readline()
                    except:
                        with open(B_Intersection_list, encoding='ISO-8859-1') as data:
                            B_HEADER = data.readline()
                # print(A_HEADER,'/',B_HEADER)
                HEADER = str(str(A_HEADER) == str(B_HEADER))
            if 'True' in HEADER: HEADER = 'Same'
            if 'False' in HEADER: HEADER = 'Diff'

            # 4. STAN_COUNT
            STAN_COUNT = '-'
            # 파일이 csv 일 경우
            if A_Intersection_list.endswith('.csv') == True and B_Intersection_list.endswith('.csv') == True:
                try:
                    with open(A_Intersection_list) as data:
                        reader = csv.reader(data)
                        STAN_COUNT = format(sum(1 for row in data), ',')
                except:
                    with open(A_Intersection_list, encoding='utf-8') as data:
                        reader = csv.reader(data)
                        STAN_COUNT = format(sum(1 for row in data), ',')
            # 파일이 xlsx 일경우
            if A_Intersection_list.endswith('.xlsx') == True and B_Intersection_list.endswith('.xlsx') == True:
                A = openpyxl.load_workbook(A_Intersection_list).active
                STAN_COUNT = format(A.max_row, ',')
            # 파일이 xls 일경우
            if A_Intersection_list.endswith('.xls') == True and B_Intersection_list.endswith('.xls') == True:
                try:
                    A = openpyxl.load_workbook(A_Intersection_list).active
                    STAN_COUNT = format(A.max_row, ',')
                except:
                    A = xlrd.open_workbook(A_Intersection_list)
                    AS = A.sheet_by_index(0)
                    STAN_COUNT = format(AS.ncols, ',')
            # shp 일 경우 dbf파일의 헤더를 비교
            if A_Intersection_list.endswith('.shp') == True and B_Intersection_list.endswith('.shp') == True:
                try:
                    AA = A_Intersection_list.replace('.shp', '.dbf')
                    STAN_COUNT = format(DBF(AA).header.numrecords, ',')
                except:
                    AA = A_Intersection_list.replace('.shp', '.dbf')
                    STAN_COUNT = format(DBF(AA, encoding='utf-8').header.numrecords, ',')
            # txt 파일 일 경우
            if A_Intersection_list.endswith('.txt') == True and B_Intersection_list.endswith('.txt') == True:
                try:
                    with open(A_Intersection_list) as data:
                        lines = data.read().splitlines()
                except:
                    try:
                        with open(A_Intersection_list, encoding='utf-8') as data:
                            lines = data.read().splitlines()
                    except:
                        with open(A_Intersection_list, encoding='ISO-8859-1') as data:
                            lines = data.read().splitlines()
                STAN_COUNT = format(len(lines), ',')
            # text 파일 일 경우
            if A_Intersection_list.endswith('.text') == True and B_Intersection_list.endswith('.text') == True:
                try:
                    with open(A_Intersection_list) as data:
                        lines = data.read().splitlines()
                except:
                    try:
                        with open(A_Intersection_list, encoding='utf-8') as data:
                            lines = data.read().splitlines()
                    except:
                        with open(A_Intersection_list, encoding='ISO-8859-1') as data:
                            lines = data.read().splitlines()
                STAN_COUNT = format(len(lines), ',')

            # 5. COMPARE_COUNT
            COMPARE_COUNT = '-'
            # 파일이 csv 일 경우
            if A_Intersection_list.endswith('.csv') == True and B_Intersection_list.endswith('.csv') == True:
                try:
                    with open(B_Intersection_list) as data:
                        reader = csv.reader(data)
                        COMPARE_COUNT = format(sum(1 for row in data), ',')
                except:
                    with open(B_Intersection_list, encoding='utf-8') as data:
                        reader = csv.reader(data)
                        COMPARE_COUNT = format(sum(1 for row in data), ',')
            # 파일이 xlsx 일경우
            if A_Intersection_list.endswith('.xlsx') == True and B_Intersection_list.endswith('.xlsx') == True:
                try:
                    B = openpyxl.load_workbook(B_Intersection_list).active
                    COMPARE_COUNT = format(B.max_row, ',')
                except:
                    B = xlrd.open_workbook(B_Intersection_list)
                    BS = B.sheet_by_index(0)
                    COMPARE_COUNT = format(BS.ncols, ',')
            # 파일이 xls 일경우
            if A_Intersection_list.endswith('.xls') == True and B_Intersection_list.endswith('.xls') == True:
                try:
                    B = openpyxl.load_workbook(B_Intersection_list).active
                    COMPARE_COUNT = format(B.max_row, ',')
                except:
                    B = xlrd.open_workbook(B_Intersection_list)
                    BS = B.sheet_by_index(0)
                    COMPARE_COUNT = format(BS.ncols, ',')
            # shp 일 경우 dbf파일의 헤더를 비교
            if A_Intersection_list.endswith('.shp') == True and B_Intersection_list.endswith('.shp') == True:
                try:
                    BB = B_Intersection_list.replace('.shp', '.dbf')
                    COMPARE_COUNT = format(DBF(BB).header.numrecords, ',')
                except:
                    BB = B_Intersection_list.replace('.shp', '.dbf')
                    COMPARE_COUNT = format(DBF(BB, encoding='utf-8').header.numrecords, ',')
            # txt 파일 일 경우
            if A_Intersection_list.endswith('.txt') == True and B_Intersection_list.endswith('.txt') == True:
                try:
                    with open(B_Intersection_list) as data:
                        lines = data.read().splitlines()
                except:
                    try:
                        with open(B_Intersection_list, encoding='utf-8') as data:
                            lines = data.read().splitlines()
                    except:
                        with open(B_Intersection_list, encoding='ISO-8859-1') as data:
                            lines = data.read().splitlines()
                COMPARE_COUNT = format(len(lines), ',')
            # text 파일 일 경우
            if A_Intersection_list.endswith('.text') == True and B_Intersection_list.endswith('.text') == True:
                try:
                    with open(B_Intersection_list) as data:
                        lines = data.read().splitlines()
                except:
                    try:
                        with open(B_Intersection_list, encoding='utf-8') as data:
                            lines = data.read().splitlines()
                    except:
                        with open(B_Intersection_list, encoding='ISO-8859-1') as data:
                            lines = data.read().splitlines()
                COMPARE_COUNT = format(len(lines), ',')
            # 6. DIFF_COUNT 계산
            try:
                DIFF_COUNT = format(int(COMPARE_COUNT.replace(',', '')) - int(STAN_COUNT.replace(',', '')), ',')
            except:
                DIFF_COUNT = '-'

            # 7. 인코딩 비교
            ENCODING = '-'
            A_ENCODING = ''
            B_ENCODING = ''
            # 파일이 csv 일경우
            if A_Intersection_list.endswith('.csv') == True and B_Intersection_list.endswith('.csv') == True:
                with open(A_Intersection_list, 'rb') as data:
                    A_ENCODING = chardet.detect(data.read(1000))
                    A_ENCODING = A_ENCODING['encoding']
                with open(B_Intersection_list, 'rb') as data:
                    B_ENCODING = chardet.detect(data.read(1000))
                    B_ENCODING = B_ENCODING['encoding']
                # print(A_ENCODING,'//',B_ENCODING)
                ENCODING = str(A_ENCODING == B_ENCODING)
            # 파일이 xlsx 일경우
            if A_Intersection_list.endswith('.xlsx') == True and B_Intersection_list.endswith('.xlsx') == True:
                with open(A_Intersection_list, 'rb') as data:
                    A_ENCODING = chardet.detect(data.read(1000))
                    A_ENCODING = A_ENCODING['encoding']
                with open(B_Intersection_list, 'rb') as data:
                    B_ENCODING = chardet.detect(data.read(1000))
                    B_ENCODING = B_ENCODING['encoding']
                # print(A_ENCODING,'//',B_ENCODING)
                ENCODING = str(A_ENCODING == B_ENCODING)
            # 파일이 xls 일경우
            if A_Intersection_list.endswith('.xls') == True and B_Intersection_list.endswith('.xls') == True:
                with open(A_Intersection_list, 'rb') as data:
                    A_ENCODING = chardet.detect(data.read(1000))
                    A_ENCODING = A_ENCODING['encoding']
                with open(B_Intersection_list, 'rb') as data:
                    B_ENCODING = chardet.detect(data.read(1000))
                    B_ENCODING = B_ENCODING['encoding']
                #             # print(A_ENCODING,'//',B_ENCODING)
                ENCODING = str(A_ENCODING == B_ENCODING)
            # shp 일 경우 dbf파일
            if A_Intersection_list.endswith('.shp') == True and B_Intersection_list.endswith('.shp') == True:
                try:
                    AA = A_Intersection_list.replace('.shp', '.dbf')
                    A_ENCODING = DBF(AA).encoding
                except:
                    AA = A_Intersection_list.replace('.shp', '.dbf')
                    A_ENCODING = DBF(AA, encoding='utf-8').encoding
                try:
                    BB = B_Intersection_list.replace('.shp', '.dbf')
                    B_ENCODING = DBF(BB).encoding
                except:
                    BB = B_Intersection_list.replace('.shp', '.dbf')
                    B_ENCODING = DBF(BB, encoding='utf-8').encoding
                # print(A_ENCODING, '//', B_ENCODING)
                ENCODING = str(A_ENCODING == B_ENCODING)
            # txt 파일 일 경우
            if A_Intersection_list.endswith('.txt') == True and B_Intersection_list.endswith('.txt') == True:
                with open(A_Intersection_list, 'rb') as data:
                    A_ENCODING = chardet.detect(data.read(1000))
                    A_ENCODING = A_ENCODING['encoding']
                with open(B_Intersection_list, 'rb') as data:
                    B_ENCODING = chardet.detect(data.read(1000))
                    B_ENCODING = B_ENCODING['encoding']
                # print(A_ENCODING,'//',B_ENCODING)
                ENCODING = str(A_ENCODING == B_ENCODING)
            # text 파일인 경우
            if A_Intersection_list.endswith('.text') == True and B_Intersection_list.endswith('.text') == True:
                with open(A_Intersection_list, 'rb') as data:
                    A_ENCODING = chardet.detect(data.read(1000))
                    A_ENCODING = A_ENCODING['encoding']
                with open(B_Intersection_list, 'rb') as data:
                    B_ENCODING = chardet.detect(data.read(1000))
                    B_ENCODING = B_ENCODING['encoding']
                # print(A_ENCODING,'//',B_ENCODING)
                ENCODING = str(A_ENCODING == B_ENCODING)
            if 'True' in ENCODING: ENCODING = 'Same'
            if 'False' in ENCODING: ENCODING = 'Diff' + '//' + A_ENCODING + '//' + B_ENCODING

            # 7. 구분자 비교
            DELIMITER = '-'
            A_ENCODING_INFO = A_ENCODING
            if A_ENCODING == 'ascii': A_ENCODING_INFO = None
            if A_ENCODING == 'UTF-8-SIG': A_ENCODING_INFO = 'UTF-8'
            if A_ENCODING == '': A_ENCODING_INFO = None
            if A_ENCODING == 'EUC-KR': A_ENCODING_INFO = None
            B_ENCODING_INFO = B_ENCODING
            if B_ENCODING == 'ascii': B_ENCODING_INFO = None
            if B_ENCODING == 'UTF-8-SIG': B_ENCODING_INFO = 'UTF-8'
            if B_ENCODING == '': B_ENCODING_INFO = None
            if B_ENCODING == 'EUC-KR': B_ENCODING_INFO = None
            # 파일이 csv 일경우
            if A_Intersection_list.endswith('.csv') == True and B_Intersection_list.endswith('.csv') == True:
                with open(A_Intersection_list, encoding=A_ENCODING_INFO) as data:
                    A = data.readline()
                    A_DELIMITER = detect(A)
                with open(B_Intersection_list, encoding=B_ENCODING_INFO) as data:
                    B = data.readline()
                    B_DELIMITER = detect(B)
                # print(A_DELIMITER,'//',B_DELIMITER)
                DELIMITER = str(A_DELIMITER == B_DELIMITER)
            # txt 파일 일 경우
            if A_Intersection_list.endswith('.txt') == True and B_Intersection_list.endswith('.txt') == True:
                with open(A_Intersection_list, encoding=A_ENCODING_INFO) as data:
                    A = data.readline()
                    A_DELIMITER = detect(A)
                with open(B_Intersection_list, encoding=B_ENCODING_INFO) as data:
                    B = data.readline()
                    B_DELIMITER = detect(B)
                # print(A_DELIMITER,'/',B_DELIMITER)
                DELIMITER = str(A_DELIMITER == B_DELIMITER)
            # text 파일 일 경우
            if A_Intersection_list.endswith('.text') == True and B_Intersection_list.endswith('.text') == True:
                with open(A_Intersection_list, encoding=A_ENCODING_INFO) as data:
                    A = data.readline()
                    A_DELIMITER = detect(A)
                with open(B_Intersection_list, encoding=B_ENCODING_INFO) as data:
                    B = data.readline()
                    B_DELIMITER = detect(B)
                # print(A_DELIMITER,'/',B_DELIMITER)
                DELIMITER = str(A_DELIMITER == B_DELIMITER)
            if 'True' in DELIMITER: DELIMITER = 'Same'
            if 'False' in DELIMITER: DELIMITER = 'Diff' + '//' + A_DELIMITER + '//' + B_DELIMITER
            sheet.append([FILE_NAME, A_Intersection_list_capacity, B_Intersection_list_capacity, DIFF_CAPACITY, HEADER,
                          STAN_COUNT, COMPARE_COUNT, DIFF_COUNT, ENCODING, DELIMITER, A_Intersection_list,
                          B_Intersection_list])
            self.progressBar.setValue(count)
            QApplication.processEvents()
        date = time.strftime('%Y%m%d', time.localtime(time.time()))
        wb.save('배포데이터검수결과_' + date + '.xlsx')

        wb = openpyxl.load_workbook('배포데이터검수결과_' + date + '.xlsx')
        ws2 = wb.active
        cellref = 'A1:Z9999'  # adjusted to show more cells
        red_font = Font(size=14, bold=True, color='ffffff')
        red_fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
        ws2.conditional_formatting.add(cellref,
                                       FormulaRule(formula=[f'NOT(ISERROR(SEARCH("Diff",{cellref})))'], stopIfTrue=True,
                                                   fill=red_fill))
        wb.save('배포데이터검수결과_' + date + '.xlsx')
        print('비교 종료')
        # ws2.conditional_formatting.add(cellref, Rule(type='containsText', operator='containsText', formula=['fail'], dxf = DifferentialStyle(fill=red_fill, font=red_font)))
        # ws2.conditional_formatting.add(cellref, CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))
        return None

    def Compare_Main(self):
        start = time.time()
        stand_text = self.Standard_line.text()
        compare_text = self.Compare_line.text()

        self.file_remove()p
        self.textBrowser.append('기존 리스트 파일 제거 완료')
        self.textBrowser.append('폴더 리스트 수집 시작')
        self.stand_list_save(stand_text)
        self.textBrowser.append('기준 폴더 리스트 저장 완료')
        self.compare_list_save(compare_text)
        self.textBrowser.append('비교 폴더 리스트 저장 완료')

        self.textBrowser.append('비교 시작')
        self.detail_compare()
        self.textBrowser.append('비교 완료')
        end = time.time() - start
        self.textBrowser.append('time : '+str(end))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Folder_Compare = QtWidgets.QWidget()
    ui = Ui_Folder_Compare()
    ui.setupUi(Folder_Compare)
    Folder_Compare.show()
    sys.exit(app.exec_())
